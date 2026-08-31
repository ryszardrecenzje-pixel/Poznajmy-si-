import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Konfiguracja strony
st.set_page_config(page_title="Intymny Check-in dla Par", page_icon="❤️", layout="centered")

# W pliku app.py w miejscu, gdzie chcesz widzieć przycisk:
if st.button("📲 Zainstaluj aplikację"):
    st.info("Aby zainstalować aplikację na telefonie lub komputerze, użyj menu przeglądarki (trzy kropki w prawym górnym rogu) i wybierz 'Zainstaluj aplikację' lub 'Dodaj do ekranu głównego'.")

# Wstrzyknięcie manifestu PWA, ikony Apple oraz przycisku instalacji PWA
st.markdown(
    """
    <link rel="manifest" href="/static/manifest.json">
    <link rel="apple-touch-icon" href="https://raw.githubusercontent.com/ryszardrecenzje-pixel/Poznajmy-si-/main/static/icon.png">
    
    <script>
    let deferredPrompt;
    window.addEventListener('beforeinstallprompt', (e) => {
        e.preventDefault();
        deferredPrompt = e;
        const installBtn = document.getElementById('pwa-install-btn');
        if (installBtn) {
            installBtn.style.display = 'block';
        }
    });

    function triggerInstall() {
        if (deferredPrompt) {
            deferredPrompt.prompt();
            deferredPrompt.userChoice.then((choiceResult) => {
                if (choiceResult.outcome === 'accepted') {
                    console.log('Użytkownik zainstalował aplikację');
                }
                deferredPrompt = null;
            });
        }
    }
    </script>

    <style>
    .install-container {
        position: fixed;
        bottom: 20px;
        right: 20px;
        z-index: 99999;
    }
    #pwa-install-btn {
        background-color: #E91E63;
        color: white;
        border: none;
        padding: 10px 16px;
        border-radius: 20px;
        font-weight: bold;
        cursor: pointer;
        box-shadow: 0 4px 10px rgba(0,0,0,0.3);
        display: none;
        font-family: sans-serif;
    }
    #pwa-install-btn:hover {
        background-color: #C2185B;
    }
    </style>

    <div class="install-container">
        <button id="pwa-install-btn" onclick="triggerInstall()">📱 Dodaj skrót aplikacji</button>
    </div>
    """,
    unsafe_allow_html=True
)

# --- MOTYW CSS ---
st.markdown("""
    <style>
    .stApp {
        background-color: #FFF0F5;
        background-image: url("data:image/svg+xml,%3Csvg width='60' height='60' viewBox='0 0 60 60' xmlns='http://www.w3.org/2000/svg'%3E%3Cpath d='M30 50.8C29.6 50.6 8 36.8 3.5 28.5C-1.5 19.5 3 8 13.5 5.5C20 4 27.5 9 30 16.5C32.5 9 40 4 46.5 5.5C57 8 61.5 19.5 56.5 28.5C52 36.8 30.4 50.6 30 50.8Z' fill='%23F8BBD0' fill-opacity='0.25'/%3E%3C/svg%3E");
        background-size: 50px 50px;
        color: #4A0E2E;
    }
    h1, h2, h3 {
        color: #AD1457 !important;
        font-family: 'Segoe UI', sans-serif;
    }
    p, label, .stMarkdown, .stRadio div, .stCheckbox span {
        color: #581845 !important;
        font-size: 1.05rem;
    }
    .stButton>button {
        background: linear-gradient(90deg, #E91E63 0%, #880E4F 100%);
        color: white;
        border: none;
        border-radius: 12px;
        padding: 0.6rem 1.2rem;
        font-weight: bold;
        box-shadow: 0 4px 12px rgba(233, 30, 99, 0.3);
    }
    .stButton>button:hover {
        background: linear-gradient(90deg, #C2185B 0%, #AD1457 100%);
        color: #FFFFFF;
    }
    .stAlert, div[data-baseweb="notification"] {
        background-color: #FFFFFF !important;
        color: #4A0E2E !important;
        border: 2px solid #EC407A !important;
        border-radius: 10px;
        box-shadow: 0 4px 10px rgba(236, 64, 122, 0.15);
    }
    .stAlert p, div[data-baseweb="notification"] p {
        color: #4A0E2E !important;
        font-weight: 500;
    }
    </style>
""", unsafe_allow_html=True)

# --- INICJALIZACIJA STANU SESJI DLA ZGÓD ---
if "regulamin_zaakceptowany" not in st.session_state:
    st.session_state.regulamin_zaakceptowany = False

# =========================================================================
# EKRAN STARTOWY / PRAWNY (JEŚLI NIE ZAAKCEPTOWANO)
# =========================================================================
if not st.session_state.regulamin_zaakceptowany:
    st.title("💖 Witaj w Intymnym Check-inie dla Par")
    st.markdown("Zanim przejdziemy do wspólnego odkrywania nastrojów, potrzeb i granic, zapoznaj się z poniższymi informacjami i wyraź wymagane zgody.")

    with st.expander("📜 Regulamin Korzystania z Aplikacji oraz Disclaimer Prawny"):
        st.markdown(
            "**1. Charakter aplikacji:** Aplikacja ma charakter wyłącznie edukacyjny, relacyjny i rozrywkowy. Nie stanowi porady medycznej, psychologicznej ani terapii par.  \n"
            "**2. Odpowiedzialność:** Autorzy nie ponoszą odpowiedzialności za decyzje podjęte na podstawie wyników check-inu. W sprawach kryzysowych zalecamy kontakt z profesjonalistą.  \n"
            "**3. Wymagania:** Korzystanie z aplikacji jest przeznaczone wyłącznie dla osób pełnoletnich."
        )

    with st.expander("🔒 Polityka Prywatności i Informacja o Ciasteczkach (Cookies)"):
        st.markdown(
            "**1. Dane osobowe:** Aplikacja działa w trybie sesyjnym. Wprowadzone przez Ciebie dane (imiona, odpowiedzi) są przetwarzane wyłącznie w pamięci przeglądarki (Session State) na potrzeby bieżącego korzystania i generowania pliku Excel. Nie zapisujemy Twoich intymnych odpowiedzi na stałych zewnętrznych serwerach.  \n"
            "**2. Pliki Cookies:** Aplikacja może wykorzystywać pliki cookies sesyjne niezbędne do prawidłowego działania interfejsu Streamlit."
        )

    st.divider()
    st.subheader("Wymagane zgody i oświadczenia:")

    zgoda_wiek = st.checkbox("Oświadczam, że mam ukończone 18 lat.")
    zgoda_regulamin = st.checkbox("Przeczytałem/am i akceptuję Regulamin oraz Disclaimer prawny.")
    zgoda_prywatnosc = st.checkbox("Zapoznałem/am się z Polityką Prywatności i wyrażam zgodę na przetwarzanie danych w pamięci sesji.")

    st.markdown("")
    
    if st.button("🚀 Wejdź do aplikacji"):
        if zgoda_wiek and zgoda_regulamin and zgoda_prywatnosc:
            st.session_state.regulamin_zaakceptowany = True
            st.rerun()
        else:
            st.error("⚠️ Musisz zaznaczyć wszystkie wymagane zgody, aby przejść dalej!")

# =========================================================================
# WŁAŚCIWA APLIKACJA (PO ZAAKCEPTOWANIU ZGÓD)
# =========================================================================
else:
    st.title("💖 Przedspotkaniowy Check-in dla Par")
    st.write("Słodkie, intymne narzędzie do odkrywania nastrojów, potrzeb i granic – blisko siebie lub na odległość.")

    # Baza pytań na Icebreaker (Przełamywacz lodów)
    baza_icebreakerow = [
        "Co w naszym związku sprawia, że czujesz się najsilniej kochany/a?",
        "Gdybyśmy mogli natychmiast przenieść się w dowolne miejsce na świecie na ten weekend, gdzie by to było?",
        "Jaka moja cecha lub zachowanie najbardziej Cię urzeka, kiedy o tym pomyślisz?",
        "O jakiej naszej wspólnej chwili najczęściej myślisz z uśmiechem na twarzy?",
        "Gdyby nasz związek był książką, to jaki nosiłby tytuł?",
        "Jaka jest jedna mała rzecz, którą mogę zrobić w tym tygodniu, żeby odciążyć Cię mentalnie?",
        "Jaki był Twój ulubiony moment z nami z ubiegłego miesiąca?",
        "O jakim naszym wspólnym przeżyciu marzysz, a jeszcze tego nie zrobiliśmy?",
        "Co najbardziej zaskoczyło Cię pozytywnie we mnie od początku naszej znajomości?",
        "Gdybyśmy mieli spędzić cały jutrzejszy dzień leżąc w łóżku, o czym byśmy rozmawiali?"
    ]

    # Sekcja Icebreaker
    with st.container():
        st.info("🧊 **Przełamywacz lodów (Icebreaker na dziś):**")
        if st.button("✨ Wylosuj jedno sekretne pytanie na dziś"):
            wylosowane_pytanie = random.choice(baza_icebreakerow)
            st.success(f"💬 **Wylosowane pytanie:** {wylosowane_pytanie}")

    st.divider()

    # Wybór trybu relacji
    tryb_relacji = st.radio(
        "Wybierz tryb:",
        ["💞 Jesteśmy razem (na jednym urządzeniu / porównanie na żywo)", "✈️ Relacja na odległość (każdy wypełnia u siebie i przesyła plik)"],
        horizontal=False
    )

    st.divider()

    # Inicjalizacja stanu
    if "imie_1" not in st.session_state:
        st.session_state.imie_1 = "Osoba A"
    if "imie_2" not in st.session_state:
        st.session_state.imie_2 = "Osoba B"
    if "odpowiedzi" not in st.session_state:
        st.session_state.odpowiedzi = {}

    # Konfiguracja imion
    col1, col2 = st.columns(2)
    with col1:
        st.session_state.imie_1 = st.text_input("Imię Pierwszej Osoby:", value=st.session_state.imie_1)
    with col2:
        st.session_state.imie_2 = st.text_input("Imię Drugiej Osoby:", value=st.session_state.imie_2)

    st.divider()

    # Wybór kto teraz wypełnia
    aktualna_osoba = st.radio("Kto teraz wypełnia quiz?", [st.session_state.imie_1, st.session_state.imie_2], horizontal=True)

    st.divider()

    # --- FORMULARZ PYTAŃ ---
    st.header("0. Klimat i Lokalizacja Randki")
    miejsce_randki = st.selectbox(
        f"{aktualna_osoba}: Gdzie i w jakim klimacie spędzamy dzisiaj czas?",
        ["W domu pod kocem 🛋️", "Kolacja przy świecach 🕯️", "Na mieście / randka poza domem 🌃", "Romantyczny spacer 🌙", "Własna niespodzianka 🎁"]
    )

    st.header("1. Nastrój, Energia i Bezpieczeństwo")
    energia = st.slider(f"{aktualna_osoba}: Jaki masz poziom energii?", 1, 10, 5)
    nastroj_slowo = st.selectbox(
        f"{aktualna_osoba}: Określ swój obecny nastrój jednym słowem:",
        ["Zrelaksowany/a", "Podekscytowany/a", "Zmęczony/a", "Czule nastrojony/a", "Stresujący dzień", "Gotowy/a na wszystko"]
    )
    bezpiecz_pytanie = st.text_input(f"{aktualna_osoba}: Co mogę dzisiaj zrobić, żebyś poczuł/a się przy mnie w 100% bezpiecznie?")

    st.header("2. Emocje i Granice")
    komfort = st.text_area(f"{aktualna_osoba}: Czy jest coś o Twoim stanie emocjonalnym, o czym partner powinien wiedzieć?")
    granice = st.text_input(f"{aktualna_osoba}: Czego dzisiaj ZUPEŁNIE NIE CHCESZ / jakie są Twoje granice?")
    potrzeba = st.selectbox(
        f"{aktualna_osoba}: Czego najbardziej dzisiaj potrzebujesz?",
        ["Wspólnego wyciszenia i rozmowy", "Czułości i przytulasów", "Dobrej zabawy i śmiechu", "Bliskości fizycznej / Intymności"]
    )

    st.header("3. Seksualność, Fantazje i Oczekiwania")
    seks_ochota = st.select_slider(
        f"{aktualna_osoba}: Jak dużą masz dziś ochotę na zbliżenia?",
        options=["Brak ochoty / Tylko czułość", "Otwarty/a, jeśli powoli", "Umiarkowana", "Duża", "100% ogień! 🔥"]
    )
    pikantna_niespodzianka = st.text_input(f"{aktualna_osoba}: Na jaką małą, pikantną niespodziankę masz ochotę?")
    motyw_przewodni = st.text_input(f"{aktualna_osoba}: Gdyby ten wieczór miał motyw przewodni, to jaki?")

    st.header("4. Jestem gotowy/wa na...")
    st.write("Zaznacz aktywności, na które czujesz się dziś w pełni gotowy/a:")
    gotowosc_opcje = [
        "Długie przytulanie i bliskość emocjonalną",
        "Masaż pleców / karku",
        "Zmysłowy masaż całego ciała",
        "Gorący prysznic / kąpiel we dwoje",
        "Grę wstępną z naciskiem na dotyk",
        "Pełną intymność i zbliżenie",
        "Eksperymenty i nowe doznania"
    ]

    zaznaczone_gotowosci = []
    for opcja in gotowosc_opcje:
        if st.checkbox(opcja, key=f"{aktualna_osoba}_{opcja}"):
            zaznaczone_gotowosci.append(opcja)

    st.header("5. Życzenia i Inicjatywy (Opcjonalnie)")
    zrob_ze_mna = st.text_input(f"{aktualna_osoba}: Zrób ze mna coś... (np. porwij mnie gdzieś, zrób mi masaż, pogaskaj po włosach)")
    chce_zrobic = st.text_input(f"{aktualna_osoba}: Chcę zrobić coś... (np. chcę Ci zrobić pyszną kolację, chcę Cię dzisiaj zaskoczyć)")

    st.header("6. Pytania i Docenienie")
    komplement = st.text_input(f"{aktualna_osoba}: Jaki komplement chcesz usłyszeć dzisiaj wieczorem?")
    wlasne_pytanie = st.text_input(f"{aktualna_osoba}: Masz specjalne pytanie lub prośbę do partnera?")

    st.divider()

    # Przycisk zapisu w sesji
    if st.button(f"Zapisz odpowiedzi dla: {aktualna_osoba}"):
        st.session_state.odpowiedzi[aktualna_osoba] = {
            "Data": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "Klimat i Miejsce": miejsce_randki,
            "Energia": energia,
            "Nastrój": nastroj_slowo,
            "Bezpieczeństwo": bezpiecz_pytanie if bezpiecz_pytanie else "Brak uwag",
            "Komfort/Emocje": komfort if komfort else "Brak uwag",
            "Granice": granice if granice else "Brak ograniczeń",
            "Potrzeba": potrzeba,
            "Ochota na intymność": seks_ochota,
            "Pikantna niespodzianka": pikantna_niespodzianka if pikantna_niespodzianka else "Brak",
            "Motyw przewodni": motyw_przewodni if motyw_przewodni else "Romantyczny wieczór",
            "Gotowość": zaznaczone_gotowosci,
            "Zrób ze mna coś": zrob_ze_mna if zrob_ze_mna else "Brak",
            "Chcę zrobić coś": chce_zrobic if chce_zrobic else "Brak",
            "Komplement": komplement if komplement else "Miłe słowo",
            "Własne pytanie": wlasne_pytanie if wlasne_pytanie else "Brak"
        }
        st.success(f"Zapisano pomyślnie odpowiedzi dla {aktualna_osoba}! 🎉")

    st.divider()

    # --- FUNKCJA GENEROWANIA EXCELA (RÓŻOWO-BŁĘKITNY) ---
    def generuj_excel_karty(dane_karty, imie_wypelniajacego):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Karta Spotkania"
        ws.views.sheetView[0].showGridLines = False
        
        CELL_BG_BLUE = "E3F2FD"
        HEADER_BG_PINK = "F8BBD0"
        TEXT_DARK = "4A0E2E"
        TEXT_ACCENT = "AD1457"
        BORDER_COLOR = "90CAF9"
        
        title_font = Font(name="Segoe UI", size=14, bold=True, color="880E4F")
        header_font = Font(name="Segoe UI", size=10, bold=True, color=TEXT_DARK)
        header_fill = PatternFill(start_color=HEADER_BG_PINK, end_color=HEADER_BG_PINK, fill_type="solid")
        
        data_font = Font(name="Segoe UI", size=10, color="0D47A1")
        accent_data_font = Font(name="Segoe UI", size=10, bold=True, color=TEXT_ACCENT)
        data_fill = PatternFill(start_color=CELL_BG_BLUE, end_color=CELL_BG_BLUE, fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color=BORDER_COLOR),
            right=Side(style='thin', color=BORDER_COLOR),
            top=Side(style='thin', color=BORDER_COLOR),
            bottom=Side(style='thin', color=BORDER_COLOR)
        )
        
        ws.merge_cells('A1:B1')
        ws['A1'] = f"💖 KARTA SPOTKANIA: {imie_wypelniajacego.upper()}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 40
        
        fields = [
            ("📅 Data i Czas", dane_karty["Data"]),
            ("👤 Kto odpowiada", imie_wypelniajacego),
            ("📍 Klimat / Miejsce randki", dane_karty["Klimat i Miejsce"]),
            ("⚡ Poziom Energii", f"{dane_karty['Energia']} / 10"),
            ("✨ Nastrój", dane_karty["Nastrój"]),
            ("🛡️ Bezpieczeństwo", dane_karty["Bezpieczeństwo"]),
            ("💭 Stan emocjonalny", dane_karty["Komfort/Emocje"]),
            ("🚫 Granice", dane_karty["Granice"]),
            ("💖 Główna Potrzeba", dane_karty["Potrzeba"]),
            ("🔥 Ochota na intymność", dane_karty["Ochota na intymność"]),
            ("🎁 Pikantna niespodzianka", dane_karty["Pikantna niespodzianka"]),
            ("🌙 Motyw przewodni", dane_karty["Motyw przewodni"]),
            ("🌹 Jestem gotowy/wa na...", ", ".join(dane_karty["Gotowość"]) if dane_karty["Gotowość"] else "Czułość i bliskość"),
            ("💌 Zrób ze mną coś...", dane_karty["Zrób ze mna coś"]),
            ("🔥 Chcę zrobić coś...", dane_karty["Chcę zrobić coś"]),
            ("✨ Komplement", dane_karty["Komplement"]),
            ("💬 Pytanie do partnera", dane_karty["Własne pytanie"])
        ]
        
        for idx, (label, val) in enumerate(fields, start=3):
            ws.row_dimensions[idx].height = 30
            
            cell_lbl = ws.cell(row=idx, column=1, value=label)
            cell_lbl.font = header_font
            cell_lbl.fill = header_fill
            cell_lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cell_lbl.border = thin_border
            
            cell_val = ws.cell(row=idx, column=2, value=val)
            if "intymność" in label.lower() or "gotowy" in label.lower() or "granice" in label.lower() or "zrób ze mną" in label.lower():
                cell_val.font = accent_data_font
            else:
                cell_val.font = data_font
                
            cell_val.fill = data_fill
            cell_val.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            cell_val.border = thin_border
            
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 60
            
        output = BytesIO()
        wb.save(output)
        return output.getvalue()


    # --- OBSŁUGA POBIERANIA / PORÓWNANIA ---
    if "Jesteśmy razem" in tryb_relacji:
        st.subheader("📥 Pobierz kartę wybranej osoby")
        osoba_do_pobrania = st.selectbox("Wybierz czyją kartę pobrać:", [st.session_state.imie_1, st.session_state.imie_2], key="pobierz_razem")
        
        if st.button("Pobierz cukierkowy plik Excel"):
            if osoba_do_pobrania in st.session_state.odpowiedzi:
                dane = st.session_state.odpowiedzi[osoba_do_pobrania]
                excel_bytes = generuj_excel_karty(dane, osoba_do_pobrania)
                st.download_button(
                    label=f"Pobierz kartę dla: {osoba_do_pobrania}",
                    data=excel_bytes,
                    file_name=f"karta_spotkania_{osoba_do_pobrania.lower()}_pink.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                dane_tymczasowe = {
                    "Data": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "Klimat i Miejsce": miejsce_randki,
                    "Energia": energia, "Nastrój": nastroj_slowo,
                    "Bezpieczeństwo": bezpiecz_pytanie or "Brak uwag",
                    "Komfort/Emocje": komfort or "Brak uwag",
                    "Granice": granice or "Brak ograniczeń",
                    "Potrzeba": potrzeba, "Ochota na intymność": seks_ochota,
                    "Pikantna niespodzianka": pikantna_niespodzianka or "Brak",
                    "Motyw przewodni": motyw_przewodni or "Romantyczny wieczór",
                    "Gotowość": zaznaczone_gotowosci,
                    "Zrób ze mną coś": zrob_ze_mna or "Brak",
                    "Chcę zrobić coś": chce_zrobic or "Brak",
                    "Komplement": komplement or "Miłe słowo",
                    "Własne pytanie": wlasne_pytanie or "Brak"
                }
                excel_bytes = generuj_excel_karty(dane_tymczasowe, osoba_do_pobrania)
                st.download_button(
                    label=f"Pobierz kartę dla: {osoba_do_pobrania}",
                    data=excel_bytes,
                    file_name=f"karta_spotkania_{osoba_do_pobrania.lower()}_pink.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        if len(st.session_state.odpowiedzi) >= 2:
            st.divider()
            st.header("✨ Porównanie Waszych odpowiedzi na żywo")
            oA = st.session_state.odpowiedzi.get(st.session_state.imie_1)
            oB = st.session_state.odpowiedzi.get(st.session_state.imie_2)
            
            c1, c2 = st.columns(2)
            with c1:
                st.markdown(f"**{st.session_state.imie_1}**")
                st.write(f"Miejsce: {oA['Klimat i Miejsce']}")
                st.write(f"Chcę zrobić: {oA['Chcę zrobić coś']}")
                st.write(f"Zrób ze mną: {oA['Zrób ze mna coś']}")
            with c2:
                st.markdown(f"**{st.session_state.imie_2}**")
                st.write(f"Miejsce: {oB['Klimat i Miejsce']}")
                st.write(f"Chcę zrobić: {oB['Chcę zrobić coś']}")
                st.write(f"Zrób ze mną: {oB['Zrób ze mna coś']}")

    else:
        st.subheader("✈️ Tryb relacji na odległość")
        st.write("Wypełnij formularz jako Ty, pobierz swoją cukierkową kartę w pliku Excel i wyślij ją partnerowi.")
        
        if st.button("Pobierz moją kartę spotkania (Excel)"):
            dane_odleglosc = {
                "Data": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "Klimat i Miejsce": miejsce_randki,
                "Energia": energia, "Nastrój": nastroj_slowo,
                "Bezpieczeństwo": bezpiecz_pytanie or "Brak uwag",
                "Komfort/Emocje": komfort or "Brak uwag",
                "Granice": granice or "Brak ograniczeń",
                "Potrzeba": potrzeba, "Ochota na intymność": seks_ochota,
                "Pikantna niespodzianka": pikantna_niespodzianka or "Brak",
                "Motyw przewodni": motyw_przewodni or "Romantyczny wieczór",
                "Gotowość": zaznaczone_gotowosci,
                "Zrób ze mną coś": zrob_ze_mna or "Brak",
                "Chcę zrobić coś": chce_zrobic or "Brak",
                "Komplement": komplement or "Miłe słowo",
                "Własne pytanie": wlasne_pytanie or "Brak"
            }
            excel_bytes = generuj_excel_karty(dane_odleglosc, aktualna_osoba)
            st.download_button(
                label=f"💾 Pobierz plik dla: {aktualna_osoba}",
                data=excel_bytes,
                file_name=f"karta_spotkania_{aktualna_osoba.lower()}_odleglosc.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
